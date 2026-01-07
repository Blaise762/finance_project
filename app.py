import streamlit as st
import pymysql
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
from db_config import MYSQL_INFO, TITLE

# ===================== 极简数据库连接+数据获取 =====================
#1:连接数据库
def get_db_conn():
    try:
        conn = pymysql.connect(**MYSQL_INFO)
        return conn
    except Exception as e:
        st.error(f"数据库连接失败: {e}")
        st.stop()

#2:获取核心数据
@st.cache_data(ttl=3600)  # 缓存1小时
def get_data(time_period_type, start_date=None, end_date=None, phone_number=None):
    conn = get_db_conn()
    
    # 根据时间粒度构建查询条件
    if time_period_type == '年度':
        # 获取当前选中年份的所有数据
        year = start_date[:4]
        where_clause = f"b.record_date LIKE '{year}%%' AND b.phone_number = %s"
    elif time_period_type == '季度':
        # 获取当前选中季度的所有数据
        year = start_date[:4]
        month = int(start_date[5:7])
        quarter = (month - 1) // 3 + 1
        if quarter == 1:
            where_clause = f"b.record_date BETWEEN '{year}-01-01' AND '{year}-03-31' AND b.phone_number = %s"
        elif quarter == 2:
            where_clause = f"b.record_date BETWEEN '{year}-04-01' AND '{year}-06-30' AND b.phone_number = %s"
        elif quarter == 3:
            where_clause = f"b.record_date BETWEEN '{year}-07-01' AND '{year}-09-30' AND b.phone_number = %s"
        else:
            where_clause = f"b.record_date BETWEEN '{year}-10-01' AND '{year}-12-31' AND b.phone_number = %s"
    elif time_period_type == '月度':
        # 获取当前选中月份的所有数据
        month = start_date[:7]
        where_clause = f"b.record_date LIKE '{month}%%' AND b.phone_number = %s"
    else:  # 自定义
        where_clause = f"b.record_date BETWEEN '{start_date}' AND '{end_date}' AND b.phone_number = %s"
    
    # 查询明细数据
    df_detail = pd.read_sql(f"""
        SELECT s.subject_name, s.subject_type, COALESCE(b.current_balance, 0) AS current_balance, b.remark, b.record_date
        FROM t_personal_balance b
        LEFT JOIN t_personal_subject s ON b.subject_id = s.subject_id
        WHERE {where_clause}
        ORDER BY b.record_date DESC
    """, conn, params=(phone_number,))
    
    # 查汇总数据（总资产/总负债/净资产）
    df_sum = pd.read_sql(f"""
        SELECT
            COALESCE(SUM(CASE WHEN s.subject_type='资产' THEN b.current_balance ELSE 0 END), 0) AS 总资产,
            COALESCE(SUM(CASE WHEN s.subject_type='负债' THEN b.current_balance ELSE 0 END), 0) AS 总负债,
            COALESCE(SUM(CASE WHEN s.subject_type='资产' THEN b.current_balance ELSE 0 END) -
            SUM(CASE WHEN s.subject_type='负债' THEN b.current_balance ELSE 0 END), 0) AS 净资产
        FROM t_personal_balance b
        LEFT JOIN t_personal_subject s ON b.subject_id = s.subject_id
        WHERE {where_clause}
    """, conn, params=(phone_number,))
    
    conn.close()
    
    # 确保数据完整性
    if df_sum.empty:
        # 如果没有数据，返回默认值
        df_sum_default = pd.Series({'总资产': 0, '总负债': 0, '净资产': 0})
        return df_detail, df_sum_default
    else:
        # 处理可能的None值，确保数值类型正确
        df_sum_filled = df_sum.iloc[0].fillna(0)
        return df_detail, df_sum_filled

#3:获取趋势数据（近3个时间单位）
@st.cache_data(ttl=3600)  # 缓存1小时
def get_trend_data(time_period_type, current_start_date, phone_number=None):
    conn = get_db_conn()
    trend_data = []
    
    # 根据时间粒度计算近3个时间单位的范围
    if time_period_type == '年度':
        current_year = int(current_start_date[:4])
        # 计算近3年的年份（包括当前年）
        years = [current_year - 2, current_year - 1, current_year]
        for year in years:
            where_clause = f"b.record_date LIKE '{year}%%' AND b.phone_number = %s"
            df = pd.read_sql(f"""
                SELECT
                    '{year}' AS period,
                    COALESCE(SUM(CASE WHEN s.subject_type='资产' THEN b.current_balance ELSE 0 END), 0) AS 总资产,
                    COALESCE(SUM(CASE WHEN s.subject_type='负债' THEN b.current_balance ELSE 0 END), 0) AS 总负债
                FROM t_personal_balance b
                LEFT JOIN t_personal_subject s ON b.subject_id = s.subject_id
                WHERE {where_clause}
            """, conn, params=(phone_number,))
            if not df.empty:
                trend_data.append(df.iloc[0])
    
    elif time_period_type == '季度':
        current_year = int(current_start_date[:4])
        current_month = int(current_start_date[5:7])
        current_quarter = (current_month - 1) // 3 + 1
        
        # 计算近3个季度的开始和结束日期
        quarters = []
        for i in range(2, -1, -1):
            q_ago = current_quarter - i
            if q_ago <= 0:
                quarter_year = current_year - 1
                quarter_num = q_ago + 4
            else:
                quarter_year = current_year
                quarter_num = q_ago
            
            if quarter_num == 1:
                q_start = f"{quarter_year}-01-01"
                q_end = f"{quarter_year}-03-31"
                period_label = f"{quarter_year}Q{quarter_num}"
            elif quarter_num == 2:
                q_start = f"{quarter_year}-04-01"
                q_end = f"{quarter_year}-06-30"
                period_label = f"{quarter_year}Q{quarter_num}"
            elif quarter_num == 3:
                q_start = f"{quarter_year}-07-01"
                q_end = f"{quarter_year}-09-30"
                period_label = f"{quarter_year}Q{quarter_num}"
            else:
                q_start = f"{quarter_year}-10-01"
                q_end = f"{quarter_year}-12-31"
                period_label = f"{quarter_year}Q{quarter_num}"
            
            quarters.append((period_label, q_start, q_end))
        
        # 按时间顺序查询数据
        for period_label, q_start, q_end in quarters:
            where_clause = f"b.record_date BETWEEN '{q_start}' AND '{q_end}' AND b.phone_number = %s"
            df = pd.read_sql(f"""
                SELECT
                    '{period_label}' AS period,
                    COALESCE(SUM(CASE WHEN s.subject_type='资产' THEN b.current_balance ELSE 0 END), 0) AS 总资产,
                    COALESCE(SUM(CASE WHEN s.subject_type='负债' THEN b.current_balance ELSE 0 END), 0) AS 总负债
                FROM t_personal_balance b
                LEFT JOIN t_personal_subject s ON b.subject_id = s.subject_id
                WHERE {where_clause}
            """, conn, params=(phone_number,))
            if not df.empty:
                trend_data.append(df.iloc[0])
    
    elif time_period_type == '月度':
        current_year = int(current_start_date[:4])
        current_month = int(current_start_date[5:7])
        
        # 计算近3个月的年月
        months = []
        for i in range(2, -1, -1):
            m_ago = current_month - i
            if m_ago <= 0:
                month_year = current_year - 1
                month_num = m_ago + 12
            else:
                month_year = current_year
                month_num = m_ago
            
            month_str = f"{month_year}-{month_num:02d}"
            period_label = month_str
            months.append((period_label, month_str))
        
        # 按时间顺序查询数据
        for period_label, month_str in months:
            where_clause = f"b.record_date LIKE '{month_str}%%' AND b.phone_number = %s"
            df = pd.read_sql(f"""
                SELECT
                    '{period_label}' AS period,
                    COALESCE(SUM(CASE WHEN s.subject_type='资产' THEN b.current_balance ELSE 0 END), 0) AS 总资产,
                    COALESCE(SUM(CASE WHEN s.subject_type='负债' THEN b.current_balance ELSE 0 END), 0) AS 总负债
                FROM t_personal_balance b
                LEFT JOIN t_personal_subject s ON b.subject_id = s.subject_id
                WHERE {where_clause}
            """, conn, params=(phone_number,))
            if not df.empty:
                trend_data.append(df.iloc[0])
    
    conn.close()
    
    # 转换为DataFrame
    if trend_data:
        trend_df = pd.DataFrame(trend_data)
        return trend_df
    else:
        # 如果没有数据，返回空DataFrame
        return pd.DataFrame(columns=['period', '总资产', '总负债'])

# ===================== 数据导入功能 =====================
# 获取所有科目信息
@st.cache_data(ttl=3600)
def get_all_subjects():
    conn = get_db_conn()
    df = pd.read_sql("SELECT subject_id, subject_name, subject_type FROM t_personal_subject ORDER BY subject_type, subject_id", conn)
    conn.close()
    return df

# 生成Excel模板
@st.cache_data(ttl=3600)
def generate_excel_template():
    # 获取当前日期和月份
    current_date = datetime.now()
    current_month = current_date.strftime("%Y-%m")
    default_date = f"{current_month}-01"
    
    # 获取所有科目
    subjects_df = get_all_subjects()
    
    # 创建模板数据
    template_data = {
        '日期': [default_date] * len(subjects_df),
        '科目名称': subjects_df['subject_name'].tolist(),
        '科目类型': subjects_df['subject_type'].tolist(),
        '金额': [0.0] * len(subjects_df),
        '备注': [''] * len(subjects_df)
    }
    
    template_df = pd.DataFrame(template_data)
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "资产负债数据"
    
    # 写入表头
    headers = ['日期', '科目名称', '科目类型', '金额', '备注']
    ws.append(headers)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # 写入数据
    for row in dataframe_to_rows(template_df, index=False, header=False):
        ws.append(row)
    
    # 移除下拉列表验证，允许用户自由输入自定义科目
    
    # 重新定义subject_names用于设置列格式
    subject_names = subjects_df['subject_name'].tolist()
    
    # 日期列格式化为YYYY-MM-DD
    for cell in ws[f"A2:A{len(subject_names)+1}"]:
        cell[0].number_format = "yyyy-mm-dd"
    
    # 金额列格式化为数字
    for cell in ws[f"D2:D{len(subject_names)+1}"]:
        cell[0].number_format = "#,##0.00"
    
    # 为科目类型列添加下拉选择（资产/负债）
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 设置数据验证规则
    dv = DataValidation(type="list", formula1='"资产,负债"', allow_blank=False)
    
    # 应用到科目类型列（C列）
    dv.sqref = "C2:C1000"  # 直接设置范围
    ws.add_data_validation(dv)
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# 解析上传的Excel文件
def parse_uploaded_file(uploaded_file):
    try:
        # 读取Excel文件
        df = pd.read_excel(uploaded_file, sheet_name=0)
        
        # 验证必要的列是否存在
        required_columns = ['日期', '科目名称', '科目类型', '金额']
        if not all(col in df.columns for col in required_columns):
            st.error(f"上传的文件缺少必要的列: {', '.join(required_columns)}")
            return None, None, None
        
        # 处理缺失值
        # 日期列不能为空
        if df['日期'].isnull().any():
            st.error("日期列不能包含空值")
            return None, None, None
        
        # 验证日期格式
        try:
            df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d')
        except:
            st.error("日期格式不正确，请使用YYYY-MM-DD格式")
            return None, None, None
        
        # 科目名称列不能为空
        if df['科目名称'].isnull().any():
            st.error("科目名称列不能包含空值")
            return None, None, None
        
        # 科目类型列不能为空
        if df['科目类型'].isnull().any():
            st.error("科目类型列不能包含空值")
            return None, None, None
        
        # 验证科目类型值
        valid_types = ['资产', '负债']
        if not df['科目类型'].isin(valid_types).all():
            st.error("科目类型必须为'资产'或'负债'")
            return None, None, None
        
        # 金额列不能为空
        if df['金额'].isnull().any():
            st.error("金额列不能包含空值")
            return None, None, None
        
        # 验证金额格式
        try:
            df['金额'] = pd.to_numeric(df['金额'])
        except:
            st.error("金额格式不正确，请输入数字")
            return None, None, None
        
        # 处理备注列（如果不存在则添加）
        if '备注' not in df.columns:
            df['备注'] = ''
        else:
            # 将备注列的NaN值替换为空字符串
            df['备注'] = df['备注'].fillna('')
        
        # 过滤掉金额为0的行
        df = df[df['金额'] != 0].copy()
        
        # 获取科目映射（名称到ID列表）
        subjects_df = get_all_subjects()
        
        # 创建科目名称到ID列表的映射
        subject_map = {}
        for _, row in subjects_df.iterrows():
            name = row['subject_name']
            if name not in subject_map:
                subject_map[name] = []
            subject_map[name].append(row['subject_id'])
        
        # 添加科目ID列，处理相同科目名称的情况
        df['subject_id'] = None
        
        # 为每个相同科目名称的行分配不同的ID
        for name, ids in subject_map.items():
            # 找出所有使用这个科目名称的行
            name_rows = df[df['科目名称'] == name]
            if not name_rows.empty:
                # 循环分配ID
                for i, (idx, row) in enumerate(name_rows.iterrows()):
                    df.at[idx, 'subject_id'] = ids[i % len(ids)]
        
        # 分离已知科目和未知科目
        known_subjects_df = df[df['subject_id'].notnull()].copy()
        unknown_subjects_df = df[df['subject_id'].isnull()].copy()
        
        # 获取未知科目列表
        unknown_subjects = unknown_subjects_df['科目名称'].unique().tolist()
        
        # 如果有已知科目，转换数据类型
        if not known_subjects_df.empty:
            known_subjects_df['subject_id'] = known_subjects_df['subject_id'].astype(int)
        
        return known_subjects_df, unknown_subjects, df
    except Exception as e:
        st.error(f"文件解析失败: {e}")
        return None, None, None

# 将数据导入到数据库
def import_data_to_db(df, phone_number):
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        
        # 开始事务
        conn.begin()
        
        # 先确保用户存在于t_user表中
        check_user_sql = "SELECT phone_number FROM t_user WHERE phone_number = %s"
        cursor.execute(check_user_sql, (phone_number,))
        user_exists = cursor.fetchone()
        
        if not user_exists:
            # 如果用户不存在，插入用户记录
            insert_user_sql = "INSERT INTO t_user (phone_number) VALUES (%s)"
            cursor.execute(insert_user_sql, (phone_number,))
        
        # 准备插入/更新语句
        sql = """
        INSERT INTO t_personal_balance (phone_number, subject_id, record_date, current_balance, remark)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            current_balance = VALUES(current_balance),
            remark = VALUES(remark)
        """
        
        # 遍历数据框并执行插入/更新
        for index, row in df.iterrows():
            cursor.execute(sql, (
                phone_number,
                row['subject_id'],
                row['日期'],
                row['金额'],
                row['备注']
            ))
        
        # 提交事务
        conn.commit()
        
        # 关闭连接
        cursor.close()
        conn.close()
        
        return True, f"成功导入 {len(df)} 条记录"
    except Exception as e:
        # 回滚事务
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        return False, f"导入失败: {str(e)}"

# ===================== Streamlit可视化 =====================
# 1. 网页基础设置
st.set_page_config(page_title=TITLE, page_icon="💰", layout="wide")

# 自定义标题样式：调小字体并改为深蓝色
st.markdown("""
<style>
/* 标题样式 - 使用更具体的选择器覆盖Streamlit默认样式 */
h1, .stHeadingContainer h1, [data-testid="stMarkdownContainer"] h1 {
    font-size: 30px !important; 
    color: #1a5276 !important; 
}

h2, .stHeadingContainer h2, [data-testid="stMarkdownContainer"] h2 {
    font-size: 22px !important; 
    color: #1a5276 !important; 
}

h3, .stHeadingContainer h3, [data-testid="stMarkdownContainer"] h3 {
    font-size: 22px !important; 
    color: #1a5276 !important; 
}

/* 直接定位Streamlit生成的指标组件,为其添加边框 */
[data-testid="metric-container"] {{ 
    padding: 1rem !important; 
    border-radius: 0.5rem !important; 
    border: 1px solid #e0e0e0 !important; 
    background-color: white !important; 
    width: 100% !important; 
    box-sizing: border-box !important; 
    margin: 0 !important; 
}}

/* 手机号输入容器样式 */
.phone-input-container {{ 
    border: 1px solid #e0e0e0; 
    border-radius: 5px; 
    padding: 20px; 
    margin: 10px 0; 
}}

/* 确保在移动端正常显示 */
@media (max-width: 768px) {
    /* 进一步调整标题大小，解决重合问题 */
    h1 {{ font-size: 22px !important; line-height: 1.2 !important; }}
    h2 {{ font-size: 18px !important; line-height: 1.2 !important; }}
    h3 {{ font-size: 16px !important; line-height: 1.2 !important; }}
    
    /* 调整指标容器 */
    [data-testid="metric-container"] {{ 
        padding: 0.4rem !important; 
    }}
    
    /* 调整自定义卡片 */
    [id^="metric-card"] {{ 
        padding: 0.4rem !important; 
    }}
    
    /* 调整图表容器 */
    .plot-container {{ 
        margin: 0 !important; 
        padding: 0 !important; 
    }}
    
    /* 调整表格样式 */
    .dataframe-container {{ 
        font-size: 12px !important; 
    }}
    
    /* 调整侧边栏 */
    [data-testid="stSidebar"] {{ 
        width: 100% !important; 
    }}
    
    /* 调整主内容区 - 减少顶部留白 */
    [data-testid="stAppViewBlockContainer"] {{ 
        padding: 0.5rem !important; 
    }}
    
    /* 调整自定义指标卡片内的字体大小 */
    .metric-card-value {{ 
        font-size: 25px !important; 
        font-weight: bold; 
    }}
    
    .metric-card-label {{ 
        font-size: 25px !important; 
    }}
}

/* 减少页面整体顶部留白 */
[data-testid="stAppViewBlockContainer"] {
    padding-top: 0rem !important;
}

/* 减少标题顶部margin */
h1 {
    margin-top: 0 !important;
    padding-top: 0 !important;
}

/* 减少Streamlit页面顶部的额外留白 */
[data-testid="stHeader"] {
    height: 0rem !important;
    padding: 0 !important;
}

/* 进一步调整页面顶部边距 */
body {
    margin-top: 0 !important;
    padding-top: 0 !important;
}
</style>
""", unsafe_allow_html=True)

# 使用markdown显示标题，并添加内联样式确保生效
st.markdown(f"<h1 style='font-size: 30px !important; color: #1a5276 !important;'>{TITLE}</h1>", unsafe_allow_html=True)

# 用户识别界面
if 'phone_number' not in st.session_state:
    st.session_state.phone_number = ''

# 只有在用户没有输入有效的手机号时，才显示输入界面
if not (st.session_state.phone_number and len(st.session_state.phone_number) == 11):
    # 创建一个简单的表单来确保所有元素被包裹在边框内
    with st.form("phone_form", border=True):
        # 显示标题
        st.markdown("<h3 style='text-align: center;'>请输入您的手机号📱</h3>", unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            phone_input = st.text_input("手机号", value=st.session_state.phone_number, placeholder="请输入11位手机号", max_chars=11)
        
        # 检查手机号格式
        if phone_input:
            if len(phone_input) == 11 and phone_input.isdigit() and phone_input.startswith('1'):
                st.session_state.phone_number = phone_input
                st.success(f"欢迎使用，手机号：{phone_input}")
                # 刷新页面以隐藏输入界面
                st.rerun()
            else:
                st.error("请输入有效的11位手机号")
        
        # 将提交按钮居中显示
        col_submit1, col_submit2, col_submit3 = st.columns([2.7, 2, 1])
        with col_submit2:
            st.form_submit_button("提交")
    
    # 阻止继续执行，直到用户输入有效手机号
    st.stop()

# 只有在用户输入有效的手机号后，才显示后续内容
if st.session_state.phone_number and len(st.session_state.phone_number) == 11:
    # 数据导入功能
    st.markdown("<h2 style='font-size: 22px !important; color: #1a5276 !important;'>数据导入</h2>", unsafe_allow_html=True)
    
    # 下载模板按钮
    excel_template = generate_excel_template()
    st.download_button(
        label="📥 下载Excel模板",
        data=excel_template,
        file_name="资产负债表导入模板.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # 文件上传组件
    uploaded_file = st.file_uploader("📤 上传已填写的Excel文件", type=["xlsx"], key="file_uploader")
    
    # 导入按钮
    if uploaded_file is not None:
        if st.button("🚀 开始导入数据", key="import_button"):
            with st.spinner("正在解析文件..."):
                # 解析上传的文件
                known_subjects_df, unknown_subjects, full_df = parse_uploaded_file(uploaded_file)
                
                if known_subjects_df is not None:
                    all_subjects_added = True
                    new_subjects_map = {}
                    
                    # 处理所有数据，包括已知和未知科目
                    with st.spinner("正在导入数据..."):
                        # 连接数据库
                        conn = get_db_conn()
                        cursor = conn.cursor()
                        conn.begin()
                        
                        try:
                            # 处理未知科目，直接从Excel读取科目类型
                            unknown_subjects_df = full_df[full_df['subject_id'].isnull()].copy()
                            for index, row in unknown_subjects_df.iterrows():
                                subject_name = row['科目名称']
                                subject_type = row['科目类型']  # 从Excel读取科目类型
                                
                                # 插入新科目到数据库
                                insert_subject_sql = "INSERT INTO t_personal_subject (subject_name, subject_type) VALUES (%s, %s)"
                                cursor.execute(insert_subject_sql, (subject_name, subject_type))
                            
                            # 提交新科目的添加
                            conn.commit()
                            
                            # 清除缓存，确保获取最新的科目数据
                            st.cache_data.clear()
                            
                            # 重新获取所有科目映射，包括新添加的
                            subjects_df = get_all_subjects()
                            
                            # 创建科目名称到ID列表的映射（处理重复科目名称）
                            subject_map = {}
                            for _, row in subjects_df.iterrows():
                                name = row['subject_name']
                                if name not in subject_map:
                                    subject_map[name] = []
                                subject_map[name].append(row['subject_id'])
                            
                            # 为每个相同科目名称的行分配不同的ID
                            full_df['subject_id'] = None
                            for name, ids in subject_map.items():
                                # 找出所有使用这个科目名称的行
                                name_rows = full_df[full_df['科目名称'] == name]
                                if not name_rows.empty:
                                    # 循环分配ID
                                    for i, (idx, row) in enumerate(name_rows.iterrows()):
                                        full_df.at[idx, 'subject_id'] = ids[i % len(ids)]
                            
                            # 转换为整数类型
                            full_df['subject_id'] = full_df['subject_id'].astype(int)
                            
                            # 将数据导入到数据库
                            success, message = import_data_to_db(full_df, st.session_state.phone_number)
                            if success:
                                st.success(message)
                                # 清除缓存并重新加载数据
                                st.cache_data.clear()
                                st.rerun()
                            else:
                                st.error(message)
                        except Exception as e:
                            conn.rollback()
                            st.error(f"数据导入失败: {e}")
                        finally:
                            cursor.close()
                            conn.close()
    
    # 添加分隔线
    st.markdown("---")
    
    # 2. 时间选择控件
    # 使用三列布局将控件排成一行
    col1, col2, col3 = st.columns(3)

    with col1:
        time_period = st.selectbox("选择时间粒度", ["年度", "季度", "月度", "自定义"])

    # 初始化日期变量
    start_date = None
    end_date = None

    # 根据选择的时间粒度显示不同的控件
    if time_period == "年度":
        with col2:
            selected_year = st.selectbox("选择年份", [2023, 2024, 2025, 2026], index=3)  # 默认2026年
        start_date = f"{selected_year}-01-01"
        end_date = f"{selected_year}-12-31"
    elif time_period == "季度":
        with col2:
            selected_year = st.selectbox("选择年份", [2023, 2024, 2025, 2026], index=3)  # 默认2026年
        with col3:
            selected_quarter = st.selectbox("选择季度", [1, 2, 3, 4])
        if selected_quarter == 1:
            start_date = f"{selected_year}-01-01"
            end_date = f"{selected_year}-03-31"
        elif selected_quarter == 2:
            start_date = f"{selected_year}-04-01"
            end_date = f"{selected_year}-06-30"
        elif selected_quarter == 3:
            start_date = f"{selected_year}-07-01"
            end_date = f"{selected_year}-09-30"
        else:
            start_date = f"{selected_year}-10-01"
            end_date = f"{selected_year}-12-31"
    elif time_period == "月度":
        with col2:
            selected_year = st.selectbox("选择年份", [2023, 2024, 2025, 2026], index=3)  # 默认2026年
        with col3:
            selected_month = st.selectbox("选择月份", range(1, 13), index=0)  # 默认1月
        start_date = f"{selected_year}-{selected_month:02d}-01"
        if selected_month == 12:
            end_date = f"{selected_year}-{selected_month}-31"
        else:
            next_month = selected_month + 1
            end_date = f"{selected_year}-{next_month:02d}-01"  # 这里可以优化为获取当月最后一天
    else:  # 自定义
        # 设置默认结束日期为当天
        default_end_date = pd.to_datetime("today")
        # 设置默认开始日期为结束日期的前一年
        default_start_date = default_end_date - pd.DateOffset(years=1)
        
        with col2:
            start_date = st.date_input("开始日期", value=default_start_date).strftime("%Y-%m-%d")
        with col3:
            end_date = st.date_input("结束日期", value=default_end_date).strftime("%Y-%m-%d")

    # 3. 加载数据
    df_detail, df_sum = get_data(time_period, start_date, end_date, st.session_state.phone_number)

    # 4. 核心指标卡片
    c1, c2, c3 = st.columns(3)

    # 确保数值不为None，使用0代替
    total_assets = df_sum['总资产'] if df_sum['总资产'] is not None else 0
    total_liabilities = df_sum['总负债'] if df_sum['总负债'] is not None else 0
    net_assets = df_sum['净资产'] if df_sum['净资产'] is not None else 0

    # 创建自定义指标卡片函数 - 添加颜色参数控制数值颜色
    def create_metric_card(label, value, value_color="#000000"):
        return f"""
        <div style="
            padding: 1rem;
            border-radius: 0.5rem;
            border: 1px solid #e0e0e0;
            background-color: white;
            width: 100%;
            box-sizing: border-box;
            text-align: center;
        ">
            <div class="metric-card-label" style="font-size: 14px; color: #666; margin-bottom: 0.5rem;">{label}</div>
            <div class="metric-card-value" style="font-size: 24px; font-weight: bold; color: {value_color};">{value}</div>
        </div>
        """

    # 添加自定义指标卡片，设置不同数值颜色
    with c1:
        st.markdown(create_metric_card("总资产 💰", f"¥{total_assets:,.2f}", value_color="#1a5276"), unsafe_allow_html=True)  # 深蓝色

    with c2:
        st.markdown(create_metric_card("总负债 💳", f"¥{total_liabilities:,.2f}", value_color="#ff0000"), unsafe_allow_html=True)  # 红色

    with c3:
        st.markdown(create_metric_card("净资产 💎", f"¥{net_assets:,.2f}", value_color="#0368C9"), unsafe_allow_html=True)  # 浅蓝色

    # 5. 趋势折线图（近3个时间单位的总资产/负债变化）
    st.markdown("<h2 style='font-size: 22px !important; color: #1a5276 !important;'>总资产负债趋势</h2>", unsafe_allow_html=True)
    if time_period != "自定义":  # 自定义时间粒度不显示趋势图
        # 获取趋势数据
        trend_df = get_trend_data(time_period, start_date, st.session_state.phone_number)
        
        if not trend_df.empty:
            # 生成图表标题
            if time_period == "年度":
                # 提取年份并生成标题
                years = sorted(trend_df['period'].astype(int))
                title = f"{years[0]}-{years[-1]}年总资产/负债趋势"
            elif time_period == "季度":
                # 提取季度并生成标题
                quarters = sorted(trend_df['period'])
                title = f"{quarters[0]}-{quarters[-1]}总资产/负债趋势"
            else:  # 月度
                # 提取月份并生成标题
                months = sorted(trend_df['period'])
                title = f"{months[0]}～{months[-1]}月总资产/负债趋势"
            
            # 绘制折线图，设置颜色：总负债为红色
            fig = px.line(trend_df, x='period', y=['总资产', '总负债'], 
                         title=title, 
                         markers=True, 
                         labels={'value': '金额（元）', 'period': '时间','variable': ''}, 
                         color_discrete_map={'总资产': 'blue', '总负债': 'red'})
            # 设置颜色和样式
            fig.update_traces(line=dict(width=2))  # 减少线条宽度
            fig.update_layout(
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                height=350,  # 进一步降低图表高度，减少渲染复杂度
                margin=dict(l=10, r=10, t=30, b=10),  # 减少边距，压缩图表空间
                hovermode="x unified"  # 优化悬停效果，减少渲染负担
            )
            
            st.plotly_chart(fig, width='stretch', key="trend_line")
        else:
            st.info("没有足够的历史数据生成趋势图")
    else:
        st.info("自定义时间范围不支持趋势图展示")

    # 6. 饼图（资产+负债）
    c1, c2 = st.columns(2)
    # 资产饼图
    asset_df = df_detail[df_detail['subject_type']=='资产']
    c1.markdown("<h2 style='font-size: 22px !important; color: #1a5276 !important;'>资产构成占比</h2>", unsafe_allow_html=True)
    if not asset_df.empty:
        # 创建资产饼图并优化 - 显示科目名称+占比
        asset_fig = px.pie(asset_df, values="current_balance", names="subject_name", hole=0.6)  # 增大中心孔，进一步缩小饼图半径
        asset_fig.update_traces(
            textposition="outside",  # 将标签移到饼图外部
            textfont_size=10,  # 减小字体大小，避免遮挡
            textinfo="label+percent",  # 显示科目名称+占比
            hovertemplate="%{label}: ¥%{value:,.2f} (%{percent})",  # 优化悬停显示
            insidetextorientation='auto'  # 优化内部文本方向
        )
        asset_fig.update_layout(
            height=300,  
            margin=dict(l=50, r=50, t=70, b=70),  # 增加左右边距，为标签提供更多空间
            legend=dict(font=dict(size=11)),  # 减小图例字体
            hovermode="closest"  # 优化悬停效果
        )
        c1.plotly_chart(asset_fig, width='stretch', key="asset_pie")
    else:
        c1.info("当前时间范围内没有资产数据")
    # 负债饼图
    debt_df = df_detail[df_detail["subject_type"]=="负债"]
    c2.markdown("<h2 style='font-size: 22px !important; color: #1a5276 !important;'>负债构成占比</h2>", unsafe_allow_html=True)
    if not debt_df.empty:
        # 创建负债饼图并优化 - 显示科目名称+占比
        debt_fig = px.pie(debt_df, values="current_balance", names="subject_name", hole=0.6)  # 增大中心孔，进一步缩小饼图半径
        debt_fig.update_traces(
            textposition="outside",  # 将标签移到饼图外部
            textfont_size=10,  # 减小字体大小，避免遮挡
            textinfo="label+percent",  # 显示科目名称+占比
            hovertemplate="%{label}: ¥%{value:,.2f} (%{percent})",  # 优化悬停显示
            insidetextorientation='auto'  # 优化内部文本方向
        )
        debt_fig.update_layout(
            height=300,  
            margin=dict(l=50, r=50, t=70, b=70),  # 增加左右边距，为标签提供更多空间
            legend=dict(font=dict(size=11)),  # 减小图例字体
            hovermode="closest"  # 优化悬停效果
        )
        c2.plotly_chart(debt_fig, width='stretch', key="debt_pie")
    else:
        c2.info("当前时间范围内没有负债数据")

    # 7. 明细表格（一键显示，带格式化）
    st.subheader("资产负债明细")
    if not df_detail.empty:
        # 创建一个新的DataFrame来避免SettingWithCopyWarning
        df_show = df_detail[["subject_name", "subject_type", "current_balance", "remark"]].copy()
        df_show.columns = ["科目", "类型", "金额", "备注"]
        # 金额格式化
        df_show["金额"] = df_show["金额"].apply(lambda x: f"¥{x:,.2f}")
        st.dataframe(df_show, width='stretch')  # 使用新参数width='stretch'替代use_container_width
    else:
        st.info("当前时间范围内没有数据")
else:
    st.stop()